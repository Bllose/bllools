from openpyxl import load_workbook

# 基本场景为
# 有两列值， 他们有一些内容是相等的， 我们需要以一列为基准， 将另外一列相同的值与其对齐

def aligning(absoluteDir):
    # 加载工作簿
    wb = load_workbook(r'C:\Users\cwx921932\Documents\工作内容\00000退换货重构梳理\退换货字段梳理.xlsx')
    # 获取sheet页
    ws = wb['ASC_DEVICE_BOM_T']

    colA = ws['A']
    colC = ws['C']
    colD = ws['D']

    # 将C列挪至D列
    for cellC in colC:
        if cellC.row == 1:
            continue
        colD[cellC.row - 1].value = cellC.value
        cellC.value = ''

    # 将D列对应A列的值，挪到对应A列行数的C列上
    for cell in colD:
        if cell.row == 1:
            continue
        if cell.value is None:
            continue
        for cell_ori in colA:
            if cell_ori.value is None:
                continue
            try:
                if cell.value.strip().upper() == cell_ori.value.strip().upper():
                    wr_col = 'C' + str(cell_ori.row);
                    print(u'向格子 %s 写入值 %s' % (wr_col, cell.value))
                    ws[wr_col] = cell.value
            except AttributeError as e:
                print(e)

    # 清空D列
    for cell in colD:
        cell.value = ''

    wb.save(r'C:\Users\cwx921932\Documents\工作内容\00000退换货重构梳理\saved.xlsx')
    wb.close()


load_file('')
